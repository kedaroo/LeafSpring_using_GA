# utils
import math
from prettytable import PrettyTable
from geneticalgorithm import geneticalgorithm as ga
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt

# # user inputs:
# print('ENTER REQUIREMENTS:')
# load = float(input('Enter total load in kN: '))
# delta = float(input('Enter total deflection required in mm: '))
# length = float(input('Enter distance between eyes of the leaf spring in mm: '))
# stress = float(input('Enter permissible stress for spring material in MPa: '))

# # design constraints:
# print('\nENTER CONSTRAINTS FOR THE FOLLOWING PARAMETERS:')
# b1 = int(input('Lower limit of width in mm: '))
# b2 = int(input('Upper limit of width in mm: '))
# t1 = int(input('Lower limit of thickness in mm: '))
# t2 = int(input('Upper limit of thickness in mm: '))
# nF1 = int(input('Lower limit of no. of full length leaves: '))
# nF2 = int(input('Upper limit of no. of full length leaves: '))
# nG1 = int(input('Lower limit of no. of graduated leaves: '))
# nG2 = int(input('Upper limit of no. of graduated leaves: '))

# hardcoded values:
load = 3.2
delta = 50
length = 1000
stress = 350

b1 = 20
b2 = 100
t1 = 4
t2 = 20
nF1 = 1
nF2 = 4
nG1 = 2
nG2 = 10 

# assumed parameters:
# modulus of elasticity
E = 210 * pow(10, 3) # in MPa

# central bands or U-bolts which are 80 mm apart
l = 80

# parameters for solution
load = (load * pow(10, 3)) / 2 # in Newtons
length = (length - l) / 2 # in mm

# density of plain carbon steel:
density = 7800 # in kg/m^3

# ~~~FUNCTIONS~~~

# assume bearing pressure as 8 MPa
def dia_of_eye(b):
	# W = d * b * pb
	pb = 8
	d1 = load / (b * pb)

	l2 = b + 4
	# maximum bending moment on the pin
	M = (load * l2) / 4

	# section modulus
	# Z = (pi/32) * d^3
	# d2 = math.ceil(pow(M / ((math.pi / 32) * stress), 1/3)) # rearranging sigma_b = M/Z
	d2 = pow(M / ((math.pi / 32) * stress), 1/3) # rearranging sigma_b = M/Z
	d = max(d1, d2)

	return d

def length_of_leaves(d, t, nF, nG):
	
	n = nF + nG

	len_of_leaves = list()
	for i in range(1, int(nG + 1)):
		li = (((length * 2) / (n - (nF - 1))) * i) + l
		len_of_leaves.append(li)

	for i in range(int(nF - 1)):
		li = (length * 2) + l
		len_of_leaves.append(li)
		
	master_leaf = (2 * length + l) + (2 * math.pi * (d + t))
	len_of_leaves.append(master_leaf)

	return len_of_leaves

# returns value in mm
def initial_nip(t, b, n):
	# C = round((2 * load * pow(length, 3)) / (n * E * b * pow(t, 3)), 2)
	C = (2 * load * pow(length, 3)) / (n * E * b * pow(t, 3))
	return C

# load on the clip bolts required to close the gap in Newton
def load_on_clip_bolts():
	Wb = (2 * nF1 * nG1 * load) / (n1 * ((2 * nG1) + (3 * nF1)))
	return Wb

def radius_of_curvature():
	y = delta
	R = ((pow(length, 2) / y) + y) / 2
	return R

def spring_mass(b, t, ll):
	total_volume = 0
	for i in ll:
		v = b * t * i
		total_volume += v
	mass = (total_volume / pow(1000, 3)) * density
	return mass


# ~~~GENETIC ALGORITHM APPROACH~~~

def deflection(y):
	b = y[0]
	t = y[1]
	nF = y[2]
	nG = y[3]
	return (12 * load * pow(length, 3)) / (E * b * pow(t, 3) * (2 * nG + 3 * nF))

def deflection2(b, t, nG, nF):
	return (12 * load * pow(length, 3)) / (E * b * pow(t, 3) * (2 * nG + 3 * nF))

# stress in full length leaves is equal to stress in graduated leaves
# leaves are initially stressed
def stress1(x):
    
    b = x[0]
    t = x[1]
    nF = x[2]
    nG = x[3]
    d = deflection(x)
    pen = 0
    if d < delta:
        pen = 24500 + 1000 * (delta - t)
    # elif d > (delta + 2):
    # 	pen = 17500 + 1000 * (delta - t)

    # s = ((6 * load * length) / ((nF + nG) * b * pow(t, 2)))
    return ((6 * load * length) / ((nF + nG) * b * pow(t, 2))) + pen


# stress in full length leaves
# stress in full length leaves is not equal to stress in graduated leaves
# leaves are not initially stressed
def stress2(x):
	
	b = x[0]
	t = x[1]
	nF = x[2]
	nG = x[3]
	d = deflection(x)
	pen = 0
	if d < delta:
		pen = 24000 + 1000 * (delta - d)
	return (((18 * load * length) / (b * pow(t, 2) * (2 * nG + 3 * nF))) + pen)

# stress for adaquate design
def stress3(nf, ng, b, t):
	return ((6 * load * length) / ((nF + nG) * b * pow(t, 2)))

# ~~~OPTIMIZING MASS~~~
def mass(x):
	b = x[0]
	t = x[1]
	nF = x[2]
	nG = x[3]
	n = nF + nG
	d = dia_of_eye(b)
	lens = length_of_leaves(d, t, n)
	m = spring_mass(b, t, lens)
	s = stress1(x)
	deflec = deflection(x)
	pen = 0
	print(f'Stress: {s}')
	print(f'Deflection: {deflec}')
	if deflec < delta:
		pen = 400 + 1000 * (deflec - delta)

	# if deflec < delta:
		# pen = 400 + 1000 * (delta - deflec)

	return (m + pen) 


# varbound sequence = b, t, nF, nG
b = np.array([b1, b2])
t = np.array([t1, t2])
nF = np.array([nF1, nF2])
nG = np.array([nG1, nG2])
varbound = np.array([b, t, nF, nG])
vartype = np.array([['int'], ['int'], ['int'], ['int']])

model1 = ga(function = stress1, dimension = 4, variable_type_mixed = vartype, variable_boundaries = varbound)
model2 = ga(function = stress2, dimension = 4, variable_type_mixed = vartype, variable_boundaries = varbound)
model1.run()
model2.run()

solution1 = model1.output_dict
b1 = solution1.get('variable')[0]
t1 = solution1.get('variable')[1]
nF1 = solution1.get('variable')[2]
nG1 = solution1.get('variable')[3]
n1 = nF1 + nG1

solution2 = model2.output_dict
b2 = solution2.get('variable')[0]
t2 = solution2.get('variable')[1]
nF2 = solution2.get('variable')[2]
nG2 = solution2.get('variable')[3]
n2 = nF2 + nG2

d1 = dia_of_eye(b1)
d2 = dia_of_eye(b2)

ll1 = length_of_leaves(d1, t1, nF1, nG1)
ll2 = length_of_leaves(d2, t2, nF2, nG2)
# print(f'Spring 1: {ll1}')
# print(f'Spring 2: {ll2}')

# it is same for both the types of springs
ROC = radius_of_curvature()

# the following two are only applicable to initially strssed leaf spring (i.e. case 1)
nip = initial_nip(t1, b1, n1)
preload = load_on_clip_bolts()

# induced stress
convergence1 = model1.report
sigma1 = convergence1[-1]

convergence2 = model2.report
sigma2 = convergence2[-1]

if sigma1 > stress:
	print('\nInduced stress exceeds design stress. Please increase design stress or limits of the parameter dimensions')

if sigma2 > stress:
	print('\nInduced stress exceeds design stress. Please increase design stress or limits of the parameter dimensions')

mass1 = spring_mass(b1, t1, ll1)
mass2 = spring_mass(b2, t2, ll2)

# ~~~WRITING DATA TO EXCEL~~~

wb = load_workbook(filename = 'Output.xlsx')
ws = wb['Sheet1']

# wrting user requirements
ws['B2'] = load
ws['B3'] = delta
ws['B4'] = ((length * 2) + l)
ws['B5'] = stress

# induced atress
ws['B10'] = sigma1
ws['B11'] = sigma2

# width
ws['E2'] = b1
ws['F2'] = b2

# thickness
ws['E3'] = t1
ws['F3'] = t2

# no. of full length leaves
ws['E4'] = nF1 
ws['F4'] = nF2

# no. of graduated leaves
ws['E5'] = nG1
ws['F5'] = nG2

# radius of curvature
ws['E6'] = ROC
ws['F6'] = ROC

# nip
ws['E7'] = nip

# preload 
ws['E8'] = preload 

# Assumptions
ws['B14'] = E
ws['B15'] = l

# eye diameter
ws['E9'] = d1
ws['F9'] = d2

# length of leaves for case 1
for i in range(1, len(ll1) + 1):
	cell = 'E' + str(i + 9)
	ws[cell] = ll1[-i]

# length of leaves for case 2
for i in range(1, len(ll2) + 1):
	cell = 'F' + str(i + 9)
	ws[cell] = ll2[-i]

wb.save('Output.xlsx')

print('The calculated design variables have successfully been saved to Output.xlsx')
